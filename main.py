#!/usr/bin/python3
# coding: UTF-8

import openpyxl
import sys
import subprocess

# ansibleのiniファイルを定義(置換前)
org_file_name = "all.yml.org"
# ansibleのiniファイルを定義(置換後)
file_name = "all.yml"

# コマンドライン引数の数を確認
if len(sys.argv) != 2:
    print("input error")
    sys.exit(1)

# Excelファイル名を変数に代入
args = sys.argv
target = args[1]

wb = openpyxl.load_workbook(target)
sheet = wb.get_sheet_by_name('Sheet1')
ws = wb.worksheets[0]


# osuserの数/データを取得
values_osuser = []
osuser_list = []
for cell in ws[8]:
        values_osuser.append(cell.value)
osuser = filter(None, values_osuser)
osuser_list = list(osuser)

# パスワードの数/データを取得
values_passwd = []
passwd_list = []
for cell in ws[9]:
        values_passwd.append(cell.value)
passwd = filter(None, values_passwd)
passwd_list = list(passwd)

osuser_count = 0
for pwd in passwd_list:
    if pwd == passwd_list[0]:
        continue
    osuser_count += 1
    #print(pwd + " " + osuser_list[osuser_count])
    with open('passwd.txt', mode='a') as f:
        f.write(pwd + " " + osuser_list[osuser_count] + "\n")

# userの数/データを取得
values_user = []
user_list = []
for cell in ws[16]:
        values_user.append(cell.value)
user = filter(None, values_user)
user_list = list(user)

# vhostsの数/データを取得
values_vhosts = []
vhosts_list = []
for cell in ws[14]:
        values_vhosts.append(cell.value)
vhosts = filter(None, values_vhosts)
vhosts_list = list(vhosts)

user_count = 0
for vh in vhosts_list:
    if vh == vhosts_list[0]:
        continue
    user_count += 1
    #print(vh + " " + user_list[user_count])
    with open('vhosts.txt', mode='a') as f:
        f.write(vh + " " + user_list[user_count] + "\n")

# phpのデータを取得
values_php = []
php_list = []
for cell in ws[21]:
        values_php.append(cell.value)
php = filter(None, values_php)
php_list = list(php)

for pl in php_list:
    if pl == php_list[0]:
        continue
    #print(pl)
    with open('php.txt', mode='a') as f:
        f.write(pl + "\n")

##################################################
# Excel内データの取得関数
def get_cell(x, y):
    param = sheet.cell(row=x,column=y).value
    return param

# 必要なセルの情報を変数に代入
php_ver = str(get_cell(20, 2))
root_pass = get_cell(9, 2)

# 置換前のiniファイルを開く
with open(org_file_name, encoding="cp932") as f:
    data_lines = f.read()

# 置換
data_lines = data_lines.replace("xxx", php_ver)
data_lines = data_lines.replace("yyy", root_pass)

# 置換後のiniファイルを作成
with open(file_name, 'w', encoding="cp932") as f:
    f.write(data_lines)
##################################################

subprocess.call(["shell/user_add.sh", "passwd.txt"])
subprocess.call(["shell/vhosts_add.sh", "vhosts.txt"])
subprocess.call(["shell/php_add.sh", "php.txt"])
subprocess.call(["rm", "-f", "passwd.txt"])
subprocess.call(["rm", "-f", "vhosts.txt"])
subprocess.call(["rm", "-f", "php.txt"])
